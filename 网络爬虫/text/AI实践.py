from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from bs4 import BeautifulSoup
import json
import time
from datetime import datetime


def get_edge_driver():
    """统一创建Edge浏览器驱动"""
    opt = webdriver.EdgeOptions()
    opt.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Edge(options=opt)
# ==================== 任务1：新浪新闻（selenium + openpyxl）====================
def sina():
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "时间", "简介", "链接", "来源"])
    for p in range(1, 6):
        api_url = (
            "https://feed.mix.sina.com.cn/api/roll/get"
            f"?pageid=153&lid=2516&num=20&page={p}"
        )
        driver.get(api_url)
        time.sleep(1.5)
        body_text = driver.find_element(By.TAG_NAME, "body").text
        data = json.loads(body_text)
        for item in data["result"]["data"]:
            title = item.get("title", "")
            ctime = item.get("ctime", "")
            dt = datetime.fromtimestamp(int(ctime)) if ctime else ""
            intro = (item.get("intro", "") or "")[:100]
            url = item.get("url", "")
            ws.append([title, str(dt), intro, url, "新浪新闻"])
        print(f"新浪第{p}页: {len(data['result']['data'])} 条")
    wb.save("新浪新闻.xlsx")
    driver.quit()
# ==================== 任务2：人邮图书（selenium + bs4 + openpyxl）====================
def books():
    import re
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["书名", "作者", "ISBN", "定价", "详情链接"])
    for p in range(1, 6):
        driver.get(f"https://www.ptpress.com.cn/publishing/book/#page={p}")
        time.sleep(6)  # 等待Nuxt.js完全渲染
        soup = BeautifulSoup(driver.page_source, "html.parser")
        cards = soup.select(".el-col-6")  # Element UI栅格，每本书一张卡片
        if not cards:
            print(f"图书第{p}页: 未找到数据")
            break
        count = 0
        for card in cards:
            try:
                title_tag = card.select_one("p[title]")
                title = title_tag.get("title", "").strip() if title_tag else ""
                if not title:
                    continue
                p_tags = card.select("p")
                author = p_tags[1].text.strip() if len(p_tags) > 1 else ""
                img_tag = card.select_one("img")
                isbn = ""
                if img_tag:
                    src = img_tag.get("src", "")
                    m = re.search(r"(\d{3}-\d-\d{2,4}-\d{4,6}-\d)", src)
                    isbn = m.group(1) if m else ""
                a_tag = card.select_one("a")
                href = a_tag.get("href", "") if a_tag else ""
                link = "https://www.ptpress.com.cn" + href if href.startswith("/") else href
                price = "见详情页"
                ws.append([title, author, isbn, price, link])
                count += 1
            except Exception:
                continue
        print(f"图书第{p}页: {count} 本")
    wb.save("人邮图书.xlsx")
    driver.quit()
# ==================== 任务3：自选网站——彼岸壁纸（selenium + openpyxl）====================
def pic():
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["图片标题", "图片链接", "详情页", "分类", "页码"])
    for p in range(1, 6):
        if p == 1:
            url = "https://pic.netbian.com/"
        else:
            url = f"https://pic.netbian.com/index_{p}.html"
        driver.get(url)
        time.sleep(2)
        # 正确的选择器：每张图是一个<li>，在<ul class="clearfix">下
        items = driver.find_elements(By.CSS_SELECTOR, "ul.clearfix li")
        for li in items:
            try:
                img_tag = li.find_element(By.TAG_NAME, "img")
                a_tag = li.find_element(By.TAG_NAME, "a")
                title = img_tag.get_attribute("alt")      # 标题在alt属性中
                src = img_tag.get_attribute("src")         # 图片地址
                href = a_tag.get_attribute("href")         # 详情页链接
                if title and src:
                    ws.append([title, src, href, "壁纸", f"第{p}页"])
            except Exception:
                continue
        print(f"壁纸第{p}页")
    wb.save("图片数据.xlsx")
    driver.quit()
if __name__ == "__main__":
    sina()
    books()
    pic()

