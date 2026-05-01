from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from bs4 import BeautifulSoup
import json, time
from datetime import datetime

def get_edge_driver():
    opt = webdriver.EdgeOptions()
    opt.add_argument("--headless")
    opt.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Edge(options=opt)

# ==================== 任务1：新浪新闻 ====================
def Xl():
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "时间", "简介", "链接", "来源"])

    for p in range(1, 6):
        api = f"https://feed.mix.sina.com.cn/api/roll/get?pageid=153&lid=2516&num=20&page={p}"
        driver.get(api)
        time.sleep(1.5)
        data = json.loads(driver.find_element(By.TAG_NAME, "body").text)

        for item in data["result"]["data"]:
            title = item.get("title", "")
            dt = datetime.fromtimestamp(int(item.get("ctime", 0))) if item.get("ctime") else ""
            ws.append([title, str(dt), item.get("intro", "") or "", item.get("url", ""), "新浪新闻"])

        print(f"新浪第{p}页: {len(data['result']['data'])} 条")

    wb.save("新浪新闻.xlsx")
    driver.quit()


# ==================== 任务2：人邮图书 ====================
def ry():
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["书名", "作者", "ISBN", "定价", "详情链接"])

    for p in range(1, 6):
        driver.get(f"https://www.ptpress.com.cn/publishing/book/#page={p}")
        time.sleep(6)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        cards = soup.select(".el-col-6")

        count = 0
        for card in cards:
            try:
                title = (card.select_one("p[title]") or {}).get("title", "")
                if not title:
                    continue
                p_tags = card.select("p")
                author = p_tags[1].text.strip() if len(p_tags) > 1 else ""
                img = card.select_one("img")
                src = img.get("src", "") if img else ""
                idx = src.find("Material/")
                isbn = src[idx + 9:].split("/")[0] if idx >= 0 else ""
                a = card.select_one("a")
                href = (a or {}).get("href", "")
                link = "https://www.ptpress.com.cn" + href if href.startswith("/") else href
                ws.append([title, author, isbn, "见详情页", link])
                count += 1
            except Exception:
                continue

        print(f"图书第{p}页: {count} 本")

    wb.save("人邮图书.xlsx")
    driver.quit()


# ==================== 任务3：彼岸壁纸 ====================
def ba():
    driver = get_edge_driver()
    wb = Workbook()
    ws = wb.active
    ws.append(["图片标题", "图片链接", "详情页", "分类", "页码"])

    for p in range(1, 6):
        url = "https://pic.netbian.com/" if p == 1 else f"https://pic.netbian.com/index_{p}.html"
        driver.get(url)
        time.sleep(2)

        for li in driver.find_elements(By.CSS_SELECTOR, "ul.clearfix li"):
            try:
                img = li.find_element(By.TAG_NAME, "img")
                title, src = img.get_attribute("alt"), img.get_attribute("src")
                href = li.find_element(By.TAG_NAME, "a").get_attribute("href")
                if title and src:
                    ws.append([title, src, href, "壁纸", f"第{p}页"])
            except Exception:
                continue

        print(f"壁纸第{p}页: 完成")

    wb.save("图片数据.xlsx")
    driver.quit()
    
if __name__ == "__main__":
    Xl()
    ry()
    ba()
