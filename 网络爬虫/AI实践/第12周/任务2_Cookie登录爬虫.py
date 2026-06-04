# -*- coding: utf-8 -*-
"""
任务二：Cookie登录爬虫 ── 豆瓣读书 (book.douban.com)
技术栈：requests + http.cookiejar + selenium + openpyxl
使用前：浏览器登录豆瓣 → F12 → Application → Cookies → 复制完整 Cookie 字符串
       粘贴到下方 COOKIE_STR 变量中
"""
import re, os, time
import requests
from http import cookiejar  # Python 内置库，用于 Cookie 持久化管理
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== 配置区 ====================
# ← 请在浏览器登录豆瓣后，复制完整 Cookie 字符串替换下面内容
COOKIE_STR = (
    "dbcl2=295202229:xDOneurXNVcֵ; ck=LqZj; bid=enGuFC4Pjfo"
    # ... 请粘贴你的完整 Cookie！
)
BOOK_TAG  = "小说"       # 图书标签（小说/历史/编程/科幻…）
PAGES     = 5            # 爬取页数（≥5，每页约20本）
# ================================================

# ────── 1. 解析 Cookie 字符串 → 字典 ──────
print("=" * 50)
print("  任务二：Cookie 登录 ── 豆瓣读书")
print("=" * 50)

cookies_dict = {}
for item in COOKIE_STR.split(";"):
    if "=" in (item := item.strip()):
        k, v = item.split("=", 1)
        # 去除 latin-1 范围外的字符（HTTP Cookie 不允许非 latin-1 字符）
        v_clean = v.strip().encode("latin-1", errors="ignore").decode("latin-1")
        cookies_dict[k.strip()] = v_clean
print(f"\n[1] 解析到 {len(cookies_dict)} 个 Cookie")

# ────── 2. 使用 http.cookiejar 持久化 Cookie 到本地文件 ──────
COOKIE_FILE = "douban_cookies.txt"
jar = cookiejar.LWPCookieJar(COOKIE_FILE)

for k, v in cookies_dict.items():
    jar.set_cookie(cookiejar.Cookie(
        version=0, name=k, value=v, port=None, port_specified=False,
        domain=".douban.com", domain_specified=True, domain_initial_dot=True,
        path="/", path_specified=True, secure=False, expires=None,
        discard=False, comment=None, comment_url=None, rest={}
    ))
# LWPCookieJar.save() 在中文 Windows 下默认用 GBK 编码，Cookie 含特殊字符会挂
# 改用手动写入 UTF-8 编码的文件
with open(COOKIE_FILE, "w", encoding="utf-8") as f:
    f.write(jar.as_lwp_str(ignore_discard=True, ignore_expires=True))
print(f"    Cookie 已持久化: {os.path.abspath(COOKIE_FILE)}")

# ────── 3. requests 携带 Cookie 验证登录状态 ──────
ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
s = requests.Session()
for k, v in cookies_dict.items():
    s.cookies.set(k, v, domain=".douban.com", path="/")

print("[2] 验证 Cookie 登录状态...")
try:
    r = s.get("https://www.douban.com/", headers={"User-Agent": ua}, timeout=10)
    print("   ✓ Cookie 有效！" if r.status_code == 200 else f"   ? 状态码 {r.status_code}")
except Exception as e:
    print(f"   ? 验证异常: {e}")

# ────── 4. Selenium 注入 Cookie → 爬取图书列表 ──────
print("[3] 启动 Selenium，注入 Cookie...")
opts = webdriver.ChromeOptions()
opts.add_argument("--headless")
opts.add_argument("--window-size=1920,1080")
driver = webdriver.Chrome(options=opts)

driver.get("https://book.douban.com/")
time.sleep(2)
for k, v in cookies_dict.items():
    try: driver.add_cookie({"name": k, "value": v, "domain": ".douban.com", "path": "/"})
    except: pass
driver.refresh()
time.sleep(1)

# ────── 5. 翻页爬取 5 个字段 ──────
print(f"[4] 爬取豆瓣「{BOOK_TAG}」标签 ({PAGES} 页)...")
all_data = []

for page in range(PAGES):
    start = page * 20
    url = f"https://book.douban.com/tag/{BOOK_TAG}?start={start}&type=T"
    print(f"    第 {page+1}/{PAGES} 页...", end=" ")
    driver.get(url)
    time.sleep(2.5)

    try: WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "subject-list")))
    except: print("超时"); continue

    for item in driver.find_elements(By.CSS_SELECTOR, "li.subject-item"):
        try:
            # 字段1: 书名
            try:
                a = item.find_element(By.CSS_SELECTOR, "h2 a")
                title = a.text.strip()
                link = a.get_attribute("href")
            except: title, link = "", ""

            # 字段2: 出版信息（作者/译者/出版社/年份/定价，用 / 分隔）
            pub_info = ""
            try: pub_info = item.find_element(By.CSS_SELECTOR, "div.pub").text.strip()
            except: pass

            parts = [p.strip() for p in pub_info.split("/")] if pub_info else []
            author  = parts[0] if len(parts) >= 1 else ""
            pub     = parts[-3] if len(parts) >= 4 else (parts[2] if len(parts) >= 3 else "")
            year    = parts[-2] if len(parts) >= 4 else ""
            price   = parts[-1] if len(parts) >= 1 and ("元" in parts[-1] or "CNY" in parts[-1]) else ""

            # 字段3: 评分
            try: rating = item.find_element(By.CSS_SELECTOR, "span.rating_nums").text.strip()
            except: rating = "暂无"

            # 字段4: 评价人数
            try: count = re.sub(r"\D", "", item.find_element(By.CSS_SELECTOR, "span.pl").text)
            except: count = ""

            # 字段5: 简介
            try:
                desc = item.find_element(By.CSS_SELECTOR, "p").text.strip()
                desc = desc[:100] + "..." if len(desc) > 100 else desc
            except: desc = ""

            if title:
                all_data.append([title, author, pub, year, price, rating, count, desc, link, page+1])
        except: continue

    print(f"累计 {len(all_data)} 条")
    time.sleep(2)

driver.quit()
print(f"\n爬取完成！共 {len(all_data)} 条")

# ────── 6. openpyxl 保存 Excel ──────
print("[5] 保存到 Excel...")
wb = Workbook(); ws = wb.active; ws.title = f"豆瓣{BOOK_TAG}"

hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

cols = ["序号", "书名", "作者", "出版社", "年份", "定价", "评分", "评价数", "简介", "链接", "页码"]
for i, h in enumerate(cols, 1):
    c = ws.cell(1, i, h); c.font = hdr_font; c.fill = hdr_fill; c.border = bdr
    c.alignment = Alignment(horizontal="center", vertical="center")

for r, row in enumerate(all_data, 2):
    for c, val in enumerate([r-1] + row, 1):
        cell = ws.cell(r, c, val)
        cell.font = Font(name="微软雅黑", size=10); cell.border = bdr
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        # 评分 ≥ 9.0 标红
        if c == 7:
            try:
                if float(val) >= 9.0: cell.font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
            except: pass

for col, w in zip("ABCDEFGHIJK", [6, 28, 14, 16, 8, 8, 6, 10, 40, 32, 6]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

out = f"豆瓣{BOOK_TAG}_爬取结果.xlsx"
wb.save(out)
print(f"✓ 已保存: {os.path.abspath(out)}  |  {len(all_data)} 条 × {PAGES} 页")