# -*- coding: utf-8 -*-
"""
任务一：表单登录爬虫 ── 古诗文网 (gushiwen.cn)
技术栈：requests + matplotlib + selenium + openpyxl
使用前：修改 USERNAME / PASSWORD 为你的古诗文网账号密码
"""
import re, os, time, io
import requests
from PIL import Image
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.common.by import By

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== 配置区 ====================
USERNAME = "18343547829"          # ← 替换为你的古诗文网账号（手机号）
PASSWORD = "1234567"       # ← 替换为你的密码
KEYWORD  = "夏"                  # 搜索关键词
PAGES    = 5                     # 爬取页数（≥5）
# ================================================

ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False

# ────── 1. Selenium 打开登录页 → 获取真实会话 Cookie ──────
print("[1] Selenium 打开登录页...")
opts = webdriver.ChromeOptions()
opts.add_argument("--headless")
opts.add_argument("--window-size=1920,1080")
driver = webdriver.Chrome(options=opts)

LOGIN_PAGE = "https://so.gushiwen.cn/user/login.aspx?from=http://so.gushiwen.cn/user/collect.aspx"
driver.get(LOGIN_PAGE)
time.sleep(2)

# 将 Selenium Cookie 同步到 requests Session
selenium_cookies = driver.get_cookies()
s = requests.Session()
for ck in selenium_cookies:
    s.cookies.set(ck["name"], ck["value"], domain=ck.get("domain", ".gushiwen.cn"), path=ck.get("path", "/"))

# ────── 2. requests 下载验证码（携带 Selenium 的真实会话 Cookie）──────
print("[2] requests 下载验证码...")
# 从页面 HTML 提取验证码图片的真实 src
cap_match = re.search(r'<img[^>]*id="imgCode"[^>]*src="([^"]+)"', driver.page_source)
cap_url = "https://so.gushiwen.cn" + cap_match.group(1) if cap_match else "https://so.gushiwen.cn/RandCode.ashx"
h = {"User-Agent": ua, "Referer": LOGIN_PAGE}
cap_resp = s.get(cap_url, headers=h)
print(f"   状态:{cap_resp.status_code}  大小:{len(cap_resp.content)}B  类型:{cap_resp.headers.get('Content-Type','?')}")

# 主方案：requests 下载；备用方案：Selenium 直接截图验证码元素
if cap_resp.status_code == 200 and len(cap_resp.content) >= 100:
    cap_img = Image.open(io.BytesIO(cap_resp.content))
else:
    print("   ⚠ 回退：Selenium 截图验证码元素...")
    cap_img = Image.open(io.BytesIO(driver.find_element(By.ID, "imgCode").screenshot_as_png))

# ────── 3. matplotlib 显示验证码 → 人工输入 ──────
print("[3] 请在弹窗中查看验证码...")
plt.figure(figsize=(4, 2))
plt.imshow(cap_img)
plt.axis("off")
plt.title("输入验证码后关闭窗口")
plt.show(block=True)
captcha = input("请输入验证码: ").strip()

# ────── 4. Selenium 填写表单并提交登录 ──────
print("[4] Selenium 提交登录...")
driver.find_element(By.ID, "email").send_keys(USERNAME)
driver.find_element(By.ID, "pwd").send_keys(PASSWORD)
driver.find_element(By.ID, "code").send_keys(captcha)
driver.find_element(By.ID, "denglu").click()
time.sleep(2)
print("   ✓ 登录完成！" if "退出" in driver.page_source else "   ? 请检查账号/密码/验证码")

# ────── 5. Selenium 翻页爬取诗文搜索结果（5 字段 × 5 页）──────
print(f"[5] 爬取「{KEYWORD}」诗文搜索结果 ({PAGES} 页)...")
all_data = []

for page in range(1, PAGES + 1):
    print(f"    第 {page}/{PAGES} 页...", end=" ")
    # 用 type=shiwen 只爬诗文，避免混入名句/古籍/作者
    url = f"https://so.gushiwen.cn/search.aspx?type=shiwen&value={KEYWORD}&page={page}"
    driver.get(url)
    time.sleep(3)

    # 诗文条目：div.zongheShiwen
    items = driver.find_elements(By.CSS_SELECTOR, "div.zongheShiwen")
    if not items:
        items = driver.find_elements(By.CSS_SELECTOR, "div.main3 div.left > div > div > div.zongheShiwen")

    if not items:
        print("→ 无结果，跳过"); continue

    for item in items:
        try:
            # 标题：.timu 里的文本
            title_el = item.find_element(By.CSS_SELECTOR, "span.timu")
            title = title_el.text.strip()
        except:
            continue
        if not title:
            continue

        # 朝代 + 作者：p.source 里的 a 标签（第一个是作者，第二个是朝代）
        dynasty, author = "", ""
        try:
            source_as = item.find_elements(By.CSS_SELECTOR, "p.source a")
            for a in source_as:
                t = a.text.strip()
                if re.search(r"[〔\[].{1,4}[代朝][〕\]]", t):
                    dynasty = t
                elif t and len(t) >= 1:
                    # 去掉图片 alt（作者头像里的同名文字）
                    author = t
        except:
            pass

        # 如果作者为空，从 textarea 里提取
        if not author or not dynasty:
            try:
                ta = item.find_element(By.CSS_SELECTOR, "textarea").get_attribute("value")
                m = re.search(r"——(.{1,4}?[代朝])[·\s]*(.{2,6})《", ta)
                if m:
                    dynasty = dynasty or m.group(1)
                    author = author or m.group(2)
            except:
                pass

        # 内容：.contson
        try:
            content = item.find_element(By.CSS_SELECTOR, "div.contson").text.strip()
            content = content[:120] + "..." if len(content) > 120 else content
        except:
            content = ""

        # 标签：从 source 文本提取，或从 catogory 获取 [唐诗][宋词] 等
        tags = ""
        try:
            source_line = item.find_element(By.CSS_SELECTOR, "p.source").text.strip()
            tags_match = re.findall(r"[〔\[【](.{1,6})[〕\]】]", source_line)
            tags = "、".join(tags_match)
        except:
            pass

        if title:
            all_data.append([title, dynasty, author, content, tags, page])

    print(f"累计 {len(all_data)} 条")
    time.sleep(1.5)

driver.quit()
print(f"\n爬取完成！共 {len(all_data)} 条")

# ────── 6. openpyxl 保存 Excel ──────
print("保存到 Excel...")
wb = Workbook(); ws = wb.active; ws.title = "古诗搜索结果"

hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_align = Alignment(horizontal="center", vertical="center")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

cols = ["序号", "标题", "朝代", "作者", "内容预览", "标签/类型", "页码"]
for i, hdr in enumerate(cols, 1):
    c = ws.cell(1, i, hdr); c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = bdr

for r, row in enumerate(all_data, 2):
    for c, val in enumerate([r - 1] + row, 1):
        cell = ws.cell(r, c, val)
        cell.font = Font(name="微软雅黑", size=10); cell.border = bdr
        cell.alignment = Alignment(vertical="center", wrap_text=True)

for col, w in zip("ABCDEFG", [6, 25, 8, 14, 45, 18, 8]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

out = "古诗搜索_爬取结果.xlsx"
wb.save(out)
print(f"✓ 已保存: {os.path.abspath(out)}  |  {len(all_data)} 条 × {PAGES} 页")